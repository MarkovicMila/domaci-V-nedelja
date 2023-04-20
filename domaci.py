import pandas as pd
import openpyxl as openpyxl
import psycopg2 as psycopg2

wb=openpyxl.load_workbook('sampledatafoodsales.xlsx')
ws=wb.active
ID=ws['A']
city=ws['D']
product=ws['F']
qty=ws['G']
UnitPrice=ws['H']
s_id=pd.Series([i.value for i in ID[1:246]])
s_city=pd.Series([i.value for i in city[1:246]])
s_product=pd.Series([i.value for i in product[1:246]])
s_qty=pd.Series([i.value for i in qty[1:246]])
s_UnitPrice=pd.Series([i.value for i in UnitPrice[1:246]])
print(s_city.head())
def formiranje_tabele():
    try:
        con=psycopg2.connect(
            database='Food',
            user='postgres',
            host='localhost',
            port='5432',
            password='itoip'
        )
        cursor=con.cursor()
        com='''CREATE TABLE FOODSALES(
        ID VARCHAR(20) PRIMARY KEY NOT NULL,
        CITY VARCHAR(15) NOT NULL,
        PRODUCT VARCHAR (20) NOT NULL,
        QTY INTEGER NOT NULL,
        UNITPRICE FLOAT NOT NULL
        );'''
        cursor.execute(com)
        print('Table created successfully!')
        con.commit()
    except(Exception,psycopg2.Error) as e:
        print('Error: ',e)
    finally:
        con.close()
        cursor.close()

def upis_u_tabelu():
    try:
        con=psycopg2.connect(
            database='Food',
            user='postgres',
            host='localhost',
            port='5432',
            password='itoip'
        )
        cursor=con.cursor()
        for i in range(len(s_id)):
            com='''INSERT INTO FOODSALES VALUES('{}','{}','{}',{},{});'''.format(s_id[i],s_city[i],s_product[i],s_qty[i],s_UnitPrice[i])
            cursor.execute(com)
        con.commit()
    except(Exception,psycopg2.Error) as e:
        print('Error: ',e)
    finally:
        con.close()
        cursor.close()

def formirati_excel():
    wb=openpyxl.Workbook()
    ws=wb.active
    ws['A1']='Min Qty'
    ws['B1']=s_qty.min()
    ws['A2']='Max Qty'
    ws['B2']=s_qty.max()
    ws['A3']='Avg Qty'
    ws['B3']=s_qty.mean()

    ws['A4']='Min UnitPrice'
    ws['B4']=s_UnitPrice.min()
    ws['A5']='Max UnitPrice'
    ws['B5']=s_UnitPrice.max()
    ws['A6']='Avg UnitPrice'
    ws['B6']=s_UnitPrice.mean()

    pom=[]
    for i in range(len(s_product)):
        if s_product[i] not in pom:
            pom.append(s_product[i])
    ws['A7']='All Products'
    ws['B7']=len(pom)
    
    pom=[]
    for i in range(len(s_city)):
        if s_city[i] not in pom:
            pom.append(s_city[i])
    ws['A8']='All City'
    ws['B8']=len(pom)

    wb.save(filename='info.xlsx')
    
# formiranje_tabele()
# upis_u_tabelu()
formirati_excel()